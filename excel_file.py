from pathlib import Path
from openpyxl import Workbook, load_workbook
import logging
logger = logging.getLogger() #main logger
  
#vrati existujici nebo vytvori novej
def getExcelFile(fileName):
  logger.info("getExcelFile(fileName):")
  if Path(fileName).exists():
    logger.debug('loading file: {0}'.format(fileName))
    wb = load_workbook(fileName)
  else:
    logger.debug('creating file: {0}'.format(fileName))
    wb = Workbook()
  ws = wb.create_sheet(index=0)
  ws.sheet_properties.tabColor = "1072BA"
  return wb
###

def generateHeader(ws):
  logger.info("generateHeader(ws):")
  header = ['FlyFrom','FlyTo','PriceCZK','DepartureTime','ArrivalTime']
  logger.debug("Header: {0}".format(header))
  ws.append(header)

  